import tkinter
from docxtpl import DocxTemplate
from openpyxl import load_workbook
import eel
import datetime
from tkinter import *
import tkinter.filedialog as fd 
import os
now = datetime.datetime.now()

month_dict = {1: "январь", 2: "февраль", 3: "март", 4: "апрель", 5: "май", 6: "июнь", 7: "июль", 8: "август", 9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"}
year = now.year
day = now.day
month = month_dict[now.month]

'''
Тут грузится список дисциплин.
Это должно работать сразу после загрузки плана!
Ну шобы список был.
'''

plan_xlsx_path = ""
oop_path = ""
disc_array = []

@eel.expose
def loadDiscLists():
    global plan_xlsx_path
    discipListGenerate = load_workbook(plan_xlsx_path)
    sheetGenerator = discipListGenerate['Лист1']

    number_of_row = sheetGenerator.max_row
    spis = []
    for i in range(number_of_row):
        a = sheetGenerator['C'+ str(i + 1)].value
        b = sheetGenerator['K'+ str(i + 1)].value
        c = sheetGenerator['O'+ str(i + 1)].value
        bIsInt = True
        try: 
            b = int(b)
            c = int(c)
            if c > 400:
                bIsInt = False
        except:
            bIsInt = False

        if a != None and bIsInt:
            spis.append(a.strip())

    f = open('Lists/list_discip.txt', 'w')
    for i in range(len(spis)):
        f.writelines(spis[i]+"\n")
    f.close()

    file = open('Lists/list_discip.txt')
    for line in file:
        string = line.replace("\n", "")
        disc_array.append(string)
    file.close()

'''
конец загрузки.
'''



list_file_num = open('Lists/list_num.txt')
num_array = []
for line in list_file_num:
    string = line.replace("\n", "")
    num_array.append(string)
list_file_num.close()

list_file_spec = open('Lists/list_spec.txt')
spec_array = []
for line in list_file_spec:
    string = line.replace("\n", "")
    spec_array.append(string)
list_file_spec.close()

list_file_naprav = open('Lists/list_naprav.txt')
naprav_array = []
for line in list_file_naprav:
    string = line.replace("\n", "")
    naprav_array.append(string)
list_file_naprav.close()

NumSpecDict = dict(zip(num_array, spec_array))
SpecNumDict = dict(zip(spec_array, num_array))

def checking_values(value):
    return value if value else 0

def getExtencion(path):
    filename, file_extension = os.path.splitext(path)
    return file_extension

def btn_click(programm_discipline, number_direction, name_direction, decryption, arr_field,
                name_sostavitel, degree, kafedra, zav_kafedra, ruk_oop, stepen_ruk_oop, cel_disciplin):

    doc = DocxTemplate("Temp/Template.docx")
    global plan_xlsx_path
    wb = load_workbook(plan_xlsx_path)
    sheet = wb['Лист1']

    number_of_row = sheet.max_row
    number_string = 0

    for i in range(number_of_row):
        if sheet['C'+ str(i + 1)].value == programm_discipline:
            number_string = i + 1
            break
    
    input_cells = ['O', 'T', 'U', 'W', 'V', 'M', 'L']
    context = dict()

    text_area = ""
    for i in range(len(arr_field)):
        if ( i == len(arr_field) - 1):
            text_area += str( i + 1 ) + ") " + arr_field[i] + ".\n"
            continue
        text_area += str( i + 1 ) + ") " + arr_field[i] + ";\n"

    for cell in input_cells:
        context[f'cell_{cell}'] = checking_values(sheet[cell + str(number_string)].value)
    
    context['SRS_les'] = int(context['cell_V']) - int(context['cell_W'])
    context['pas_type']  = 'экзамен'
    context['programm_discipline'] = programm_discipline
    context['number_direction'] = number_direction
    context['name_direction'] = name_direction
    context['decryption'] = decryption
    context['day'] = day
    context['month'] = month
    context['year'] = year
    context['text_area'] = text_area
    context['name_sostavitel'] = name_sostavitel
    context['degree'] = degree
    context['kafedra'] = kafedra
    context['zav_kafedra'] = zav_kafedra
    context['ruk_oop'] = ruk_oop
    context['stepen_ruk_oop'] = stepen_ruk_oop
    context['cel_disciplin'] = cel_disciplin
    doc.render(context)
    doc.save("CompiledPrograms/" + programm_discipline + " составленная программа.docx")


eel.init('SiteLayout')

@eel.expose
def render_doc(programm_discipline, number_direction, name_direction, decryption, arr_field,
                name_sostavitel, degree, kafedra, zav_kafedra, ruk_oop, stepen_ruk_oop, cel_disciplin):
                
    btn_click(programm_discipline, number_direction, name_direction, decryption, arr_field,
                name_sostavitel, degree, kafedra, zav_kafedra, ruk_oop, stepen_ruk_oop, cel_disciplin)

@eel.expose
def getTheNum(value):
    if value == "":
        return ""
    return SpecNumDict[value]

@eel.expose
def getTheSpec(value):
    if value == "":
        return ""
    return NumSpecDict[value]

@eel.expose
def FileChoiceExcel():
    root = Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    folder = fd.askopenfilename()
    global plan_xlsx_path
    plan_xlsx_path = folder
    ext = getExtencion(folder)
    print(folder)
    return folder

@eel.expose
def FileChoiceOOP():
    root = Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    folder = fd.askopenfilename()
    global oop_path
    oop_path = folder
    ext = getExtencion(folder)
    print(folder)
    return folder


@eel.expose
def LoadToHTML():
    for i in range(len(disc_array)):
        eel.addOption(disc_array[i])

    for i in range(len(num_array)):
        eel.addOptionToNum(num_array[i])

    for i in range(len(spec_array)):
        eel.addOptionToSpec(spec_array[i])

    for i in range(len(naprav_array)):
        eel.addOptionToNaprav(naprav_array[i])

eel.start('main.html')
