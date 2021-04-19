from tkinter import *
from docxtpl import DocxTemplate
from openpyxl import load_workbook
root = Tk()


def rnt(var):
    if var == None:
        return 0 
    return var 

def btn_click():

    doc = DocxTemplate("Temp/Template.docx")
    
    wb = load_workbook('Temp/plan.xlsx')
    sheet = wb.get_sheet_by_name('Лист1')

    spec_str = spec_str_field.get()

    row_count = sheet.max_row
    column_count = sheet.max_column

    list_disc = []
    for i in range(row_count):
        discip_names_column = sheet['C'+str(i + 1)].value
        list_disc.append(discip_names_column)
    
    
    it = 0
    number_string = 0
    while list_disc[it - 1] != ProgedDisc:
        if list_disc[it] == ProgedDisc:
            number_string = str(it + 1)       
        it += 1

    NameDiscip = sheet['C'+number_string].value
    CDis = CDis_field.get()
    SpecName = SpecName_field.get()
    ScepSurname = ScepSurname_field.get()

    #number_string = str(spec_str)
    All_aud_les = rnt(sheet['O'+number_string].value)
    pr_les  = rnt(sheet['T'+number_string].value)
    LR_les = rnt(sheet['U'+number_string].value)
    KSR_les = rnt(sheet['W'+number_string].value)
    SR_les = rnt(sheet['V'+number_string].value)
    SRS_les = int(SR_les) - int(KSR_les)
    ex_hour = rnt(sheet['M'+number_string].value)
    All_hour = rnt(sheet['L'+number_string].value)
    pas_type  = 'экзамен'

    context = { 'NameDiscip' : NameDiscip,
                'CDis' : CDis,
                'SpecName' : SpecName,
                'ScepSurname' : ScepSurname, 
                'All_aud_les' : All_aud_les, 
                'pr_les' : pr_les, 
                'LR_les' : LR_les,
                'KSR_les' : KSR_les, 
                'SRS_les' : SRS_les,
                'SR_les' : SR_les, 
                'ex_hour' : ex_hour,
                'All_hour' : All_hour,
                'pas_type' : pas_type,
    }
    
    doc.render(context)
    doc.save(NameDiscip+"_prog.docx")


root['bg'] = '#fafafa'
root.title('Составить программ')
root.wm_attributes('-alpha', 1)
root.geometry('500x500')
root.resizable(width = False, height = False)

canvas = Canvas(root, width = 500, height = 500)
canvas.pack()
frame = Frame(root , bg = 'white' , )
frame.place(relwidth = 1 , relheight = 1)

title = Label(frame, text = 'Ты хуйло', bg = 'gray', font = 40)
title.pack()

btn = Button(frame, text ='КНОПКА', bg = 'yellow', command = btn_click)
btn.pack()

NameDiscip_field = Entry(frame, bg = 'white')
NameDiscip_field.pack()

CDis_field = Entry(frame, bg = 'white')
CDis_field.pack()

SpecName_field = Entry(frame, bg = 'white')
SpecName_field.pack()

ScepSurname_field = Entry(frame, bg = 'white')
ScepSurname_field.pack()

spec_str_field = Entry(frame, bg = 'white')
spec_str_field.pack()

text = Text (frame, state=DISABLED)
text.insert(INSERT, "Имя дисциплины")
text.pack()






ProgedDisc = 'История'




root.mainloop()






