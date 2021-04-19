from docxtpl import DocxTemplate
from openpyxl import load_workbook


def checking_values(value):
    return value if value else 0

def btn_click(programm_discipline, number_direction, name_direction, decryption):

    doc = DocxTemplate("Temp/Template.docx")
    
    wb = load_workbook('Temp/plan.xlsx')
    sheet = wb.get_sheet_by_name('Лист1')

    number_of_row = sheet.max_row
    nubmer_of_column = sheet.max_column

    number_string = 0

    for i in range(number_of_row):
        if sheet['C'+ str(i + 1)].value == programm_discipline:
            number_string = i + 1
            break
    
    input_cells = ['O', 'T', 'U', 'W', 'V', 'M', 'L']
    context = dict()

    print(number_string)

    for cell in input_cells:
        context[f'cell_{cell}'] = checking_values(sheet[cell + str(number_string)].value)
    
    # совсем плохо 
    context['SRS_les'] = int(context['cell_V']) - int(context['cell_W'])
    context['pas_type']  = 'экзамен'
    context['programm_discipline'] = programm_discipline
    context['number_direction'] = number_direction
    context['name_direction'] = name_direction
    context['decryption'] = decryption

    doc.render(context)
    doc.save(programm_discipline + "_programm.docx")
